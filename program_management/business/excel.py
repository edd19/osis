##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2020 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import itertools
from collections import namedtuple, defaultdict

from django.db.models import QuerySet, Prefetch
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Style, Border, Side, Color, PatternFill, Font
from openpyxl.styles.borders import BORDER_THICK
from openpyxl.styles.colors import RED, GREEN
from openpyxl.writer.excel import save_virtual_workbook

from backoffice.settings.base import LEARNING_UNIT_PORTAL_URL
from base.models.education_group_year import EducationGroupYear
from base.models.enums.prerequisite_operator import OR, AND
from base.models.group_element_year import GroupElementYear, fetch_row_sql
from base.models.learning_unit_year import LearningUnitYear
# from base.models.prerequisite import Prerequisite
# from base.models.prerequisite_item import PrerequisiteItem
from osis_common.document.xls_build import _build_worksheet, CONTENT_KEY, HEADER_TITLES_KEY, WORKSHEET_TITLE_KEY, \
    STYLED_CELLS, STYLE_NO_GRAY
from program_management.ddd.business_types import *
from program_management.ddd.repositories import load_tree
from typing import List
from django.utils import translation
from backoffice.settings.base import LANGUAGE_CODE_EN
from program_management.ddd.domain.prerequisite import Prerequisite, PrerequisiteItem, PrerequisiteItemGroup

STYLE_BORDER_BOTTOM = Style(
    border=Border(
        bottom=Side(
            border_style=BORDER_THICK, color=Color('FF000000')
        )
    )
)
STYLE_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('D1D1D1')))
STYLE_LIGHT_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('E1E1E1')))
STYLE_LIGHTER_GRAY = Style(fill=PatternFill(patternType='solid', fgColor=Color('F1F1F1')))

STYLE_FONT_RED = Style(font=Font(color=RED))
STYLE_FONT_GREEN = Style(font=Font(color=GREEN))
FONT_HYPERLINK = Font(underline='single', color='0563C1')

HeaderLine = namedtuple('HeaderLine', ['egy_acronym', 'egy_title', 'code_header', 'title_header', 'credits_header',
                                       'block_header', 'mandatory_header'])
OfficialTextLine = namedtuple('OfficialTextLine', ['text'])
LearningUnitYearLine = namedtuple('LearningUnitYearLine', ['luy_acronym', 'luy_title'])
PrerequisiteItemLine = namedtuple(
    'PrerequisiteItemLine',
    ['text', 'operator', 'luy_acronym', 'luy_title', 'credits', 'block', 'mandatory']
)
PrerequisiteOfItemLine = namedtuple(
    'PrerequisitedItemLine',
    ['text', 'luy_acronym', 'luy_title', 'credits', 'block', 'mandatory']

)
HeaderLinePrerequisiteOf = namedtuple('HeaderLinePrerequisiteOf', ['egy_acronym', 'egy_title', 'title_header',
                                                                   'credits_header', 'block_header',
                                                                   'mandatory_header'])


class EducationGroupYearLearningUnitsPrerequisitesToExcel:
    def __init__(self, year: int, code: str):
        # self.egy = egy
        self.tree = load_tree.load_from_year_and_code(year, code)

    def _to_workbook(self):
        return generate_prerequisites_workbook(self.tree)

    def to_excel(self):
        return {
            'workbook': save_virtual_workbook(self._to_workbook()),
            'acronym': self.tree.root_node.title
        }


def generate_prerequisites_workbook(tree: 'ProgramTree') -> Workbook:
    worksheet_title = _("prerequisites-%(year)s-%(acronym)s") % {"year": tree.root_node.year,
                                                                 "acronym": tree.root_node.code}
    worksheet_title = clean_worksheet_title(worksheet_title)
    workbook = Workbook(encoding='utf-8')

    excel_lines = _build_excel_lines(tree)

    return _get_workbook(tree, excel_lines, workbook, worksheet_title, 7)


def _build_excel_lines(tree: 'ProgramTree') -> List:
    content = _first_line_content(
        HeaderLine(egy_acronym=tree.root_node.code,
                   egy_title=tree.root_node.title,
                   code_header=_('Code'),
                   title_header=_('Title'),
                   credits_header=_('Cred. rel./abs.'),
                   block_header=_('Block'),
                   mandatory_header=_('Mandatory'))
    )
# qui on des pre
    for luy in tree.get_nodes_that_have_prerequisites():
        #  Bastien dit ceci : Tu peux créer une fonction dans ton fichier.py qui produit le fichier excel, à l'endroit
        #  ou tu dois afficher le compelete title, tu fais uen fonction qui reprend les bons chamsp de ton objet DDD
        #  en fonciton de la langue de l'utilisateur
        content.append(
            LearningUnitYearLine(luy_acronym=luy.code, luy_title=complete_title_i18n(luy))
        )

        for items in luy.prerequisite.prerequisite_item_groups:
            print('ci')
            for item in items.prerequisite_items:
                prerequisite_line = _build_prerequisite_line(luy.prerequisite, item,
                                                             items.prerequisite_items)
                content.append(prerequisite_line)

        # groups_generator = itertools.groupby(prerequisite.items, key=lambda item: item.group_number)
        # for key, group_gen in groups_generator:
        #
        #     group = list(group_gen)
        #     for item in group:
        #         prerequisite_line = _build_prerequisite_line(prerequisite, item, group)
        #         content.append(prerequisite_line)

    return content


def _first_line_content(header_line: HeaderLinePrerequisiteOf):
    content = list()
    content.append(
        header_line
    )
    content.append(
        OfficialTextLine(text=_("Official"))
    )
    return content


def _build_prerequisite_line(prerequisite: Prerequisite, prerequisite_item: PrerequisiteItem, group: list):
    luy_item = None
    prerequisite_item.group_number
    text = (_("has as prerequisite") + " :") \
        if prerequisite_item.group_number == 1 and prerequisite_item.position == 1 else None
    operator = _get_operator(prerequisite, prerequisite_item)
    luy_acronym = _get_item_acronym(prerequisite_item, group)
    credits = _get_item_credits(prerequisite_item)
    block = _get_item_blocks(prerequisite_item)
    mandatory = luy_item.links[0].is_mandatory if luy_item.links else None
    return PrerequisiteItemLine(
        text=text,
        operator=operator,
        luy_acronym=luy_acronym,
        luy_title=prerequisite_item.learning_unit.luys[0].complete_title_i18n,
        credits=credits,
        block=block,
        mandatory=_("Yes") if mandatory else _("No")
    )


def _get_operator(prerequisite: Prerequisite, prerequisite_item: PrerequisiteItem):
    if prerequisite_item.group_number == 1 and prerequisite_item.position == 1:
        return None
    elif prerequisite_item.position == 1:
        return _(prerequisite.main_operator)
    return _(prerequisite.secondary_operator)


def _get_item_acronym(prerequisite_item: PrerequisiteItem, group: list):
    acronym_format = "{acronym}"
    if prerequisite_item.position == 1 and len(group) > 1:
        acronym_format = "({acronym}"
    elif prerequisite_item.position == len(group) and len(group) > 1:
        acronym_format = "{acronym})"
    return acronym_format.format(acronym=prerequisite_item.learning_unit.luys[0].acronym)


def _get_item_credits(prerequisite_item: PrerequisiteItem):
    luy_item = prerequisite_item.learning_unit.luys[0]
    return " ; ".join(
        set(["{} / {:f}".format(grp.relative_credits, luy_item.credits.to_integral_value()) for grp in luy_item.links])
    )


def _get_item_blocks(prerequisite_item: PrerequisiteItem):
    luy_item = prerequisite_item.learning_unit.luys[0]
    return " ; ".join(
        [str(grp.block) for grp in luy_item.links if grp.block]
    )


def _get_style_to_apply(excel_lines: list):
    style_to_apply_dict = defaultdict(list)
    last_luy_line_index = None
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, HeaderLine):
            style_to_apply_dict[STYLE_NO_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("B{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("C{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("D{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("E{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("F{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("G{index}".format(index=index))

        elif isinstance(row, OfficialTextLine):
            style_to_apply_dict[STYLE_BORDER_BOTTOM].append("A{index}".format(index=index))

        elif isinstance(row, LearningUnitYearLine):
            style_to_apply_dict[STYLE_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[STYLE_LIGHT_GRAY].append("B{index}".format(index=index))
            last_luy_line_index = index

        elif isinstance(row, PrerequisiteItemLine):
            if row.operator == _(OR):
                style_to_apply_dict[STYLE_FONT_RED].append("B{index}".format(index=index))
            elif row.operator == _(AND):
                style_to_apply_dict[STYLE_FONT_GREEN].append("B{index}".format(index=index))

            if (last_luy_line_index - index) % 2 == 1:
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("C{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("D{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("E{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("F{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("G{index}".format(index=index))
        elif isinstance(row, PrerequisiteOfItemLine):
            if (last_luy_line_index - index) % 2 == 1:
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("C{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("D{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("E{index}".format(index=index))
                style_to_apply_dict[STYLE_LIGHTER_GRAY].append("F{index}".format(index=index))
        if isinstance(row, HeaderLinePrerequisiteOf):
            style_to_apply_dict[STYLE_NO_GRAY].append("A{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("B{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("C{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("D{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("E{index}".format(index=index))
            style_to_apply_dict[STYLE_NO_GRAY].append("F{index}".format(index=index))
    return style_to_apply_dict


def _merge_cells(excel_lines, workbook: Workbook, end_column):
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, LearningUnitYearLine):
            worksheet.merge_cells(start_row=index, end_row=index, start_column=2, end_column=end_column)


def _add_hyperlink(excel_lines, workbook: Workbook, year):
    worksheet = workbook.worksheets[0]
    for index, row in enumerate(excel_lines, 1):
        if isinstance(row, LearningUnitYearLine):
            cell = worksheet.cell(row=index, column=1)
            cell.hyperlink = LEARNING_UNIT_PORTAL_URL.format(year=year, acronym=row.luy_acronym)
            cell.font = FONT_HYPERLINK

        if isinstance(row, PrerequisiteItemLine) or isinstance(row, PrerequisiteOfItemLine):
            column_nb = 3 if isinstance(row, PrerequisiteItemLine) else 2
            cell = worksheet.cell(row=index, column=column_nb)
            cell.hyperlink = LEARNING_UNIT_PORTAL_URL.format(year=year, acronym=row.luy_acronym.strip("()"))
            cell.font = FONT_HYPERLINK


class EducationGroupYearLearningUnitsIsPrerequisiteOfToExcel:

    def __init__(self, year: int, code: str):
        self.tree = load_tree.load_from_year_and_code(year, code)

    def _to_workbook(self):
        return generate_ue_is_prerequisite_for_workbook(self.tree)

    def to_excel(self):
        return {
            'workbook': save_virtual_workbook(self._to_workbook()),
            'acronym': '?'
        }


def generate_ue_is_prerequisite_for_workbook(tree: 'ProgramTree') -> Workbook:
    worksheet_title = _("is_prerequisite_of-%(year)s-%(acronym)s") % {"year": tree.root_node.year,
                                                                      "acronym": tree.root_node.code}
    worksheet_title = clean_worksheet_title(worksheet_title)
    workbook = Workbook()

    excel_lines = _build_excel_lines_prerequisited(tree)
    return _get_workbook(tree, excel_lines, workbook, worksheet_title, 6)


def _get_workbook(tree: 'ProgramTree',
                  excel_lines: List,
                  workbook: Workbook,
                  worksheet_title: str,
                  end_column: int) -> Workbook:
    header, *content = [tuple(line) for line in excel_lines]
    style = _get_style_to_apply(excel_lines)
    worksheet_data = {
        WORKSHEET_TITLE_KEY: worksheet_title,
        HEADER_TITLES_KEY: header,
        CONTENT_KEY: content,
        STYLED_CELLS: style
    }
    _build_worksheet(worksheet_data, workbook, 0)
    _merge_cells(excel_lines, workbook, end_column)
    _add_hyperlink(excel_lines, workbook, str(tree.root_node.year))
    return workbook


def _build_excel_lines_prerequisited(tree: 'ProgramTree') -> List:
    print('ici1')
    content = _first_line_content(HeaderLinePrerequisiteOf(egy_acronym=tree.root_node.code,
                                                           egy_title=tree.root_node.title,
                                                           title_header=_('Title'),
                                                           credits_header=_('Cred. rel./abs.'),
                                                           block_header=_('Block'),
                                                           mandatory_header=_('Mandatory'))
                                  )
    for child_node in tree.get_nodes_that_are_prerequisites():
        if child_node.is_prerequisite:
            content.append(
                LearningUnitYearLine(luy_acronym=child_node.code, luy_title=child_node.title)
            )
            first = True
            for prerequisite_node in child_node.get_is_prerequisite_of():
                if child_node.year == prerequisite_node.year:
                    prerequisite_line = _build_is_prerequisite_for_line(
                        prerequisite_node,
                        first,
                        tree
                    )
                    first = False
                    content.append(prerequisite_line)
    return content


def _build_is_prerequisite_for_line(prerequisite_node: 'NodeLearningUnitYear', first, tree: 'ProgramTree') -> PrerequisiteOfItemLine:
    text = (_("is a prerequisite of") + " :") if first else None
    first_link = tree.get_first_link_occurence_using_node(prerequisite_node)
    return PrerequisiteOfItemLine(
        text=text,
        luy_acronym=prerequisite_node.code,
        luy_title=prerequisite_node.title,
        credits=first_link.relative_credits_repr,
        block=first_link.block_repr,
        mandatory=_("Yes") if first_link.is_mandatory else _("No")
    )


def clean_worksheet_title(title: str) -> str:
    # Worksheet title is max 25 chars (31 chars with sheet number) + does not accept slash present in acronyms
    return title[:25].replace("/", "_")


def get_complete_title(luy):
    if translation.get_language() == LANGUAGE_CODE_EN:
        complete_title = luy.specific_title_en
    else:
        complete_title = luy.specific_title_fr
    # TODO : je vois pas comment faire pour récupérer ici le learning_container_year
    # if luy.learning_container_year:
    #     complete_title = ' - '.join(filter(None, [luy.learning_container_year.common_title, luy.specific_title]))
    return complete_title


def get_complete_title_english(luy):
    complete_title_english = luy.specific_title_english
    # TODO : je vois pas comment faire pour récupérer ici le learning_container_year
    # if luy.learning_container_year:
    #     complete_title_english = ' - '.join(filter(None, [
    #         luy.learning_container_year.common_title_english,
    #         luy.specific_title_english,
    #     ]))
    return complete_title_english


def complete_title_i18n(luy):
    complete_title = get_complete_title(luy)
    if translation.get_language() == LANGUAGE_CODE_EN:
        complete_title = luy.complete_title_english or complete_title
    return complete_title
